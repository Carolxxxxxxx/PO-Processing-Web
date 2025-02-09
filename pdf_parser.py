import pdfplumber
import openpyxl
import re
import os

# ========== 1. 解析 PDF，提取 PO 号、Part Number、箱数 ==========
def extract_data_from_pdf(pdf_path):
    """
    解析 PDF 提取 Part Number、订单数量 和 PO 号
    """
    with pdfplumber.open(pdf_path) as pdf:
        text = pdf.pages[0].extract_text()

        if not text:
            print("❌ 无法提取文本，可能是扫描版 PDF。")
            return None, None, None

        print("📜 PDF 解析的文本内容：\n", text)

        part_numbers = []
        ordered_quantities = []
        po_number = None

        # 1. 提取 PO 号
        po_match = re.search(r"Order Number\s*([\d-]+)", text)
        if po_match:
            po_number = po_match.group(1)

        # 2. 提取 Part Number 和 订单数量
        lines = text.split("\n")
        for i in range(len(lines)):
            match = re.search(r"(BHB\d{3,}-CLRK|BHW\d{3,}-CLRK)", lines[i])
            if match:
                part_number = match.group(1)
                ordered_match = re.findall(r"(\d{2,}\.00)", lines[i])  # 提取类似 "600.00"
                if ordered_match:
                    ordered_quantity = int(float(ordered_match[-1]))  # 获取最后一个匹配数值
                else:
                    ordered_quantity = "解析错误"
                    print(f"⚠️ 无法解析数量：{lines[i]}")

                part_numbers.append(part_number)
                ordered_quantities.append(ordered_quantity)

        return part_numbers, ordered_quantities, po_number


# ========== 2. 读取价格表 ==========
def load_price_list(price_path):
    """
    读取价格表，构建 {Part Number: (Price, Units per Case)} 字典
    """
    wb = openpyxl.load_workbook(price_path)
    ws = wb.active
    price_dict = {}

    for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过标题行
        if row[0]:
            part_number = row[0]
            price = row[2] if isinstance(row[2], (int, float)) else "N/A"
            units_per_case = 250 if "Case-250" in row[1] else 200  # 判断包装单位
            price_dict[part_number] = (price, units_per_case)

    return price_dict


# ========== 3. 填充 INVOICE ==========
def fill_invoice(template_path, output_path, part_numbers, ordered_quantities, po_number, price_list):
    """
    填充 INVOICE.xlsx：
    - B15, B16, B17... 填入 Part Number
    - C15, C16, C17... 填入 订单数量（箱数）
    - E15, E16, E17... 填入 产品数量（箱数 × 单箱装量）
    - H15, H16, H17... 填入 价格（匹配价格表）
    """
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # 1. 填充 PO 号
        ws["J9"] = po_number

        # 2. 填充 Part Number、订单数量、产品数量和价格
        start_row = 15
        for i, (part_number, ordered_quantity) in enumerate(zip(part_numbers, ordered_quantities)):
            row = start_row + i

            # 确保 ordered_quantity 解析正确
            order_quantity = int(ordered_quantity) if isinstance(ordered_quantity, int) else "N/A"

            # 获取价格和单位装箱数
            if part_number in price_list:
                price, units_per_case = price_list[part_number]
            else:
                price, units_per_case = "N/A", 250  # 默认 Case-250

            # 计算产品数量
            product_quantity = order_quantity * units_per_case if isinstance(order_quantity, int) else "N/A"

            ws[f"B{row}"] = part_number
            ws[f"C{row}"] = order_quantity  # 订单数量（箱数）
            ws[f"E{row}"] = product_quantity  # 产品数量
            ws[f"H{row}"] = price  # 价格

        wb.save(output_path)
        print(f"✅ INVOICE 生成成功：{output_path}")

    except Exception as e:
        print(f"❌ 发生错误：{e}")


# ========== 4. 填充箱单 ==========
def fill_packing_list(template_path, output_path, part_numbers, ordered_quantities, po_number):
    """
    填充箱单：
    - K11 填入 PO 号
    - ABC16, ABC17... 填入 Part Number
    - D16, D17, D18... 填入 产品数量
    - F16, F17, F18... 填入 箱数
    """
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # 1. 填充 PO 号
        ws["K11"] = po_number

        # 2. 填充 Part Number、订单数量和箱数
        start_row = 16
        for i, (part_number, ordered_quantity) in enumerate(zip(part_numbers, ordered_quantities)):
            row = start_row + i
            units_per_case = 250 if "Case-250" in part_number else 200  # 默认单位
            product_quantity = ordered_quantity * units_per_case if isinstance(ordered_quantity, int) else "解析错误"

            ws[f"ABC{row}"] = part_number
            ws[f"D{row}"] = product_quantity
            ws[f"F{row}"] = ordered_quantity

        wb.save(output_path)
        print(f"✅ 箱单生成成功：{output_path}")

    except Exception as e:
        print(f"❌ 发生错误：{e}")


# ========== 5. 主程序 ==========
if __name__ == "__main__":
    # 📌 文件路径
    pdf_path = "PO2024-00-90868(6403830).pdf"
    price_list_path = "Clark11款纸袋报价更新.xlsx"
    invoice_template = "INVOICE.xlsx"
    packing_list_template = "PACKING_LIST.xlsx"

    # 解析 PO PDF
    part_numbers, ordered_quantities, po_number = extract_data_from_pdf(pdf_path)

    if part_numbers and ordered_quantities and po_number:
        invoice_output = f"INVOICE_{po_number}.xlsx"
        packing_list_output = f"PACKING_LIST_{po_number}.xlsx"

        # 读取价格表
        price_list = load_price_list(price_list_path)

        # 生成 INVOICE
        fill_invoice(invoice_template, invoice_output, part_numbers, ordered_quantities, po_number, price_list)

        # 生成箱单
        fill_packing_list(packing_list_template, packing_list_output, part_numbers, ordered_quantities, po_number)

        print("🎉 所有文件生成完毕！")
    else:
        print("❌ 解析 PO 失败，请检查 PDF 格式！")
