import pdfplumber
import openpyxl
import re
import os

# ========== 1. è§£æ PDFï¼Œæå– PO å·ã€Part Numberã€ç®±æ•° ==========
def extract_data_from_pdf(pdf_path):
    """
    è§£æ PDF æå– Part Numberã€è®¢å•æ•°é‡ å’Œ PO å·
    """
    with pdfplumber.open(pdf_path) as pdf:
        text = pdf.pages[0].extract_text()

        if not text:
            print("âŒ æ— æ³•æå–æ–‡æœ¬ï¼Œå¯èƒ½æ˜¯æ‰«æç‰ˆ PDFã€‚")
            return None, None, None

        print("ğŸ“œ PDF è§£æçš„æ–‡æœ¬å†…å®¹ï¼š\n", text)

        part_numbers = []
        ordered_quantities = []
        po_number = None

        # 1. æå– PO å·
        po_match = re.search(r"Order Number\s*([\d-]+)", text)
        if po_match:
            po_number = po_match.group(1)

        # 2. æå– Part Number å’Œ è®¢å•æ•°é‡
        lines = text.split("\n")
        for i in range(len(lines)):
            match = re.search(r"(BHB\d{3,}-CLRK|BHW\d{3,}-CLRK)", lines[i])
            if match:
                part_number = match.group(1)
                ordered_match = re.findall(r"(\d{2,}\.00)", lines[i])  # æå–ç±»ä¼¼ "600.00"
                if ordered_match:
                    ordered_quantity = int(float(ordered_match[-1]))  # è·å–æœ€åä¸€ä¸ªåŒ¹é…æ•°å€¼
                else:
                    ordered_quantity = "è§£æé”™è¯¯"
                    print(f"âš ï¸ æ— æ³•è§£ææ•°é‡ï¼š{lines[i]}")

                part_numbers.append(part_number)
                ordered_quantities.append(ordered_quantity)

        return part_numbers, ordered_quantities, po_number


# ========== 2. è¯»å–ä»·æ ¼è¡¨ ==========
def load_price_list(price_path):
    """
    è¯»å–ä»·æ ¼è¡¨ï¼Œæ„å»º {Part Number: (Price, Units per Case, NW, GW)} å­—å…¸
    """
    wb = openpyxl.load_workbook(price_path)
    ws = wb.active
    price_dict = {}

    for row in ws.iter_rows(min_row=2, values_only=True):  # è·³è¿‡æ ‡é¢˜è¡Œ
        if row[0]:
            part_number = row[0]
            price = row[2] if isinstance(row[2], (int, float)) else "N/A"
            units_per_case = 250 if "Case-250" in row[1] else 200  # åˆ¤æ–­åŒ…è£…å•ä½
            nw = row[3] if isinstance(row[3], (int, float)) else "N/A"  # å‡€é‡ï¼ˆNWï¼‰
            gw = row[4] if isinstance(row[4], (int, float)) else "N/A"  # æ¯›é‡ï¼ˆGWï¼‰
            price_dict[part_number] = (price, units_per_case, nw, gw)

    return price_dict


# ========== 3. å¡«å…… INVOICE ==========
def fill_invoice(template_path, output_path, part_numbers, ordered_quantities, po_number, price_list):
    """
    å¡«å…… INVOICE.xlsxï¼š
    - B15, B16, B17... å¡«å…¥ Part Number
    - C15, C16, C17... å¡«å…¥ è®¢å•æ•°é‡ï¼ˆç®±æ•°ï¼‰
    - E15, E16, E17... å¡«å…¥ äº§å“æ•°é‡ï¼ˆç®±æ•° Ã— å•ç®±è£…é‡ï¼‰
    - H15, H16, H17... å¡«å…¥ ä»·æ ¼ï¼ˆåŒ¹é…ä»·æ ¼è¡¨ï¼‰
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    ws["J9"] = po_number  # å¡«å…… PO å·

    start_row = 15
    for i, (part_number, ordered_quantity) in enumerate(zip(part_numbers, ordered_quantities)):
        row = start_row + i
        price, units_per_case, _, _ = price_list.get(part_number, ("N/A", 250, "N/A", "N/A"))

        product_quantity = ordered_quantity * units_per_case if isinstance(ordered_quantity, int) else "N/A"

        ws[f"B{row}"] = part_number
        ws[f"C{row}"] = ordered_quantity
        ws[f"E{row}"] = product_quantity
        ws[f"H{row}"] = price

    wb.save(output_path)
    print(f"âœ… INVOICE ç”ŸæˆæˆåŠŸï¼š{output_path}")


# ========== 4. å¡«å…… PACKING LIST ==========
def fill_packing_list(template_path, output_path, part_numbers, ordered_quantities, po_number, price_list):
    """
    å¡«å…… PACKING LIST.xlsxï¼š
    - K11 å¡«å…¥ PO å·
    - ABC17, ABC18... å¡«å…¥ Part Number
    - D17, D18, D19... **å¡«å…… Excel å…¬å¼ `=F17 * P17`**
    - F17, F18, F19... å¡«å…¥ ç®±æ•°
    - H17, H18, H19... **å¡«å…… Excel å…¬å¼ `=F17 * N17`ï¼ˆä»·æ ¼è®¡ç®—ï¼‰**
    - N17, N18, N19... **ä» `Clark11æ¬¾çº¸è¢‹æŠ¥ä»·æ›´æ–°.xlsx` æå– E åˆ—æ•°æ®ï¼ˆå•ä»·ï¼‰**
    - O17, O18, O19... **ä» `Clark11æ¬¾çº¸è¢‹æŠ¥ä»·æ›´æ–°.xlsx` æå– D åˆ—æ•°æ®ï¼ˆå‡€é‡ NWï¼‰**
    - P17, P18, P19... å¡«å…¥ ä¸€ç®±è£…å¤šå°‘åª
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    ws["K11"] = po_number  # å¡«å…… PO å·

    start_row = 17
    for i, (part_number, ordered_quantity) in enumerate(zip(part_numbers, ordered_quantities)):
        row = start_row + i
        price, units_per_case, nw, gw = price_list.get(part_number, ("N/A", 250, "N/A", "N/A"))

        ws[f"ABC{row}"] = part_number
        ws[f"D{row}"] = f"=F{row} * P{row}"  # **Excel å…¬å¼ `=F17 * P17`**
        ws[f"F{row}"] = ordered_quantity
        ws[f"H{row}"] = f"=F{row} * N{row}"  # **Excel å…¬å¼ `=F17 * N17`ï¼ˆä»·æ ¼è®¡ç®—ï¼‰**
        ws[f"N{row}"] = price  # **N åˆ—åº”æå– Clark è¡¨æ ¼ E åˆ—æ•°æ®ï¼ˆå•ä»·ï¼‰**
        ws[f"O{row}"] = nw  # **O åˆ—åº”æå– Clark è¡¨æ ¼ D åˆ—æ•°æ®ï¼ˆå‡€é‡ NWï¼‰**
        ws[f"P{row}"] = units_per_case

    wb.save(output_path)
    print(f"âœ… PACKING LIST ç”ŸæˆæˆåŠŸï¼š{output_path}")


# ========== 5. ä¸»ç¨‹åº ==========
if __name__ == "__main__":
    pdf_path = "/Users/carol/Desktop/POPICI/PO2024-00-90868(6403830).pdf"
    price_list_path = "/Users/carol/Desktop/POPICI/Clark11æ¬¾çº¸è¢‹æŠ¥ä»·æ›´æ–°.xlsx"
    invoice_template = "/Users/carol/Desktop/POPICI/INVOICE.xlsx"
    packing_list_template = "/Users/carol/Desktop/POPICI/PACKING_LIST.xlsx"

    part_numbers, ordered_quantities, po_number = extract_data_from_pdf(pdf_path)

    if part_numbers and ordered_quantities and po_number:
        price_list = load_price_list(price_list_path)
        fill_invoice(invoice_template, f"INVOICE_{po_number}.xlsx", part_numbers, ordered_quantities, po_number, price_list)
        fill_packing_list(packing_list_template, f"PACKING_LIST_{po_number}.xlsx", part_numbers, ordered_quantities, po_number, price_list)

        print("ğŸ‰ æ‰€æœ‰æ–‡ä»¶ç”Ÿæˆå®Œæ¯•ï¼")
